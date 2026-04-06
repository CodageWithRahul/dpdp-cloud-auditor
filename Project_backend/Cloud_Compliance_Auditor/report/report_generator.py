"""Utility helpers for assembling downloadable scan reports."""
import io
import math
from datetime import datetime
from typing import Iterable, Optional, Sequence

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from xml.sax.saxutils import escape

SEVERITY_ORDER = [
    ("CRITICAL", "Critical"),
    ("HIGH", "High"),
    ("MEDIUM", "Medium"),
    ("LOW", "Low"),
]

SEVERITY_COLOR_MAP = {
    "CRITICAL": colors.HexColor("#D7263D"),
    "HIGH": colors.HexColor("#F18F01"),
    "MEDIUM": colors.HexColor("#F2C057"),
    "LOW": colors.HexColor("#2EC4B6"),
}

EXCEL_SEVERITY_COLOR = {
    "CRITICAL": "FFC4C4",
    "HIGH": "FFE2BA",
    "MEDIUM": "FFF4A3",
    "LOW": "D4F1E4",
}


def _format_datetime(value: Optional[datetime]) -> str:
    if not value:
        return "N/A"
    try:
        return value.strftime("%B %d, %Y %H:%M UTC")
    except (AttributeError, ValueError):
        return str(value)


def _make_paragraph(value: str, style) -> Paragraph:
    return Paragraph(escape(str(value)), style)


def _format_scan_date_range(scan_jobs: list) -> str:
    """Format the range of scan dates that are covered in the report."""

    dates = [
        (job.completed_at or job.started_at)
        for job in scan_jobs
        if getattr(job, "started_at", None) or getattr(job, "completed_at", None)
    ]
    if not dates:
        return "N/A"

    earliest = min(dates)
    latest = max(dates)
    if earliest == latest:
        return _format_datetime(earliest)

    return f"{_format_datetime(earliest)} – {_format_datetime(latest)}"


def _format_scan_regions(scan_regions: Sequence[str] | None) -> str:
    """Return a formatted string describing the scanned regions."""

    if not scan_regions:
        return "N/A"

    if isinstance(scan_regions, str):
        return scan_regions

    return ", ".join(scan_regions)


def _build_severity_chart(severity_counts: dict) -> Optional[Drawing]:
    """Return a vertical bar chart drawing showing the severity distribution."""

    labels = [label for _, label in SEVERITY_ORDER]
    values = [severity_counts[key] for key, _ in SEVERITY_ORDER]
    if sum(values) == 0:
        return None

    max_value = max(values)
    chart_height = 110
    chart_width = 230
    chart = VerticalBarChart()
    chart.x = 10
    chart.y = 15
    chart.height = chart_height
    chart.width = chart_width
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.boxAnchor = "n"
    chart.categoryAxis.labels.angle = 0
    chart.categoryAxis.labels.dx = 0
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = max_value + max(1, max_value // 5)
    chart.valueAxis.valueStep = max(1, math.ceil(chart.valueAxis.valueMax / 5))
    chart.barWidth = 18
    chart.barLabels.nudge = 5
    chart.barLabels.angle = 90
    chart.barLabels.fontSize = 7
    chart.barLabels.boxAnchor = "n"
    chart.barLabelFormat = "%d"
    chart.strokeColor = colors.HexColor("#1F4E79")

    for index, (key, _) in enumerate(SEVERITY_ORDER):
        if index >= len(chart.bars):
            break
        chart.bars[index].fillColor = SEVERITY_COLOR_MAP.get(key, colors.grey)

    drawing = Drawing(chart_width + 40, chart_height + 40)
    drawing.add(chart)
    return drawing


def generate_pdf_report(
    findings: Iterable,
    report_context: dict,
    requested_by,
    scan_logs: Optional[Sequence] = None,
) -> io.BytesIO:
    """Return a PDF report for the provided findings."""

    buffer = io.BytesIO()
    findings_list = list(findings)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=LETTER,
        rightMargin=inch * 0.4,
        leftMargin=inch * 0.4,
        topMargin=inch * 0.5,
        bottomMargin=inch * 0.5,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=24,
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        fontSize=14,
        spaceBefore=12,
        spaceAfter=4,
    )
    body_style = styles["BodyText"]

    logo_style = ParagraphStyle(
        "Logo",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.white,
        alignment=1,
    )
    logo_table = Table(
        [[Paragraph("Cloud Auditor", logo_style)]],
        colWidths=[1.8 * inch],
        rowHeights=[0.35 * inch],
    )
    logo_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1F4E79")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    flowables = [
        logo_table,
        Spacer(1, 6),
        Paragraph("Cloud Security Scan Report", title_style),
    ]

    scan_job = report_context.get("scan_job")
    scan_jobs = report_context.get("scan_jobs") or ([scan_job] if scan_job else [])
    total_findings = len(findings_list)
    scan_date_text = _format_scan_date_range(scan_jobs)
    generation_date = timezone.now()
    generation_date_text = _format_datetime(generation_date)
    requested_by_name = requested_by.get_full_name() or requested_by.username
    requested_by_email = requested_by.email or "n/a"
    cloud_account = report_context.get("cloud_account") or getattr(scan_job, "cloud_account", None)
    provider = (
        cloud_account.get_provider_display() if cloud_account else "N/A"
    )
    account_name = cloud_account.account_name if cloud_account else "N/A"

    scan_job_ids = [str(job.id) for job in scan_jobs if job]
    scan_job_label = scan_job_ids[0] if len(scan_job_ids) == 1 else (
        f"{scan_job_ids[0]} (+{len(scan_job_ids) - 1} more)"
        if scan_job_ids
        else "N/A"
    )

    scan_scope_value = (
        "No scans" if not scan_jobs else (
            f"Scan job #{scan_job_label}" if len(scan_jobs) == 1 else f"{len(scan_jobs)} scan jobs (latest #{scan_job_label})"
        )
    )

    scan_regions_text = _format_scan_regions(report_context.get("scan_regions"))

    metadata_rows = [
        [_make_paragraph("Scan Scope:", body_style), _make_paragraph(scan_scope_value, body_style)],
        [_make_paragraph("Scan Job ID:", body_style), _make_paragraph(scan_job_label, body_style)],
        [_make_paragraph("Scan Jobs:", body_style), _make_paragraph(str(len(scan_jobs)), body_style)],
        [_make_paragraph("Total Findings:", body_style), _make_paragraph(str(total_findings), body_style)],
        [_make_paragraph("Scan Date:", body_style), _make_paragraph(scan_date_text, body_style)],
        [_make_paragraph("Cloud Account:", body_style), _make_paragraph(account_name, body_style)],
        [_make_paragraph("Provider:", body_style), _make_paragraph(provider, body_style)],
        [
            _make_paragraph("Requested By:", body_style),
            _make_paragraph(f"{requested_by_name} ({requested_by_email})", body_style),
        ],
        [_make_paragraph("Report Generated:", body_style), _make_paragraph(generation_date_text, body_style)],
        [_make_paragraph("Regions Scanned:", body_style), _make_paragraph(scan_regions_text, body_style)],
    ]

    header_table = Table(metadata_rows, hAlign="LEFT", colWidths=[1.4 * inch, 3.4 * inch])
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("INNERGRID", (0, 0), (-1, -1), 0, colors.white),
                ("BOX", (0, 0), (-1, -1), 0, colors.white),
            ]
        )
    )

    flowables.append(header_table)
    flowables.append(Spacer(1, 12))

    flowables.append(Paragraph("Severity Summary", section_style))
    severity_counts = {key: 0 for key, _ in SEVERITY_ORDER}
    for finding in findings_list:
        severity = getattr(finding, "severity", "").upper()
        if severity in severity_counts:
            severity_counts[severity] += 1

    severity_rows = [[_make_paragraph("Severity", body_style), _make_paragraph("Count", body_style)]]
    for key, label in SEVERITY_ORDER:
        severity_rows.append(
            [_make_paragraph(label, body_style), _make_paragraph(str(severity_counts[key]), body_style)]
        )

    summary_table = Table(severity_rows, colWidths=[2 * inch, 1 * inch], hAlign="LEFT")
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ]
        )
    )
    flowables.append(summary_table)
    chart = _build_severity_chart(severity_counts)
    if chart:
        flowables.append(Paragraph("Severity Distribution Chart", section_style))
        flowables.append(chart)
        flowables.append(Spacer(1, 12))
    else:
        flowables.append(Spacer(1, 12))

    flowables.append(Paragraph("Detailed Findings", section_style))

    detail_headers = [
        "Scan Job",
        "Resource ID",
        "Service",
        "Region",
        "Severity",
        "Issue Title",
        "Description",
        "Remediation",
    ]
    detail_data = [[_make_paragraph(header, body_style) for header in detail_headers]]

    if not findings_list:
        detail_data.append(
            [_make_paragraph("No findings available", body_style)] + [""] * (len(detail_headers) - 1)
        )
    else:
        for finding in findings_list:
            resource_id = finding.resource_id or "N/A"
            service = getattr(finding, "resource_type", "N/A")
            finding_job = getattr(finding, "scan_job", None)
            region = getattr(finding, "region", None)
            if not region and finding_job:
                regions_text = finding_job.target_regions_text()
                region = regions_text
            region_display = region or "N/A"
            severity = getattr(finding, "severity", "N/A")
            issue_title = getattr(finding, "issue_type", "N/A")
            description = getattr(finding, "description", "N/A")
            remediation = getattr(finding, "recommendation", "N/A")

            detail_data.append(
                [
                    _make_paragraph(str(finding_job.id) if finding_job else "N/A", body_style),
                    _make_paragraph(resource_id, body_style),
                    _make_paragraph(service, body_style),
                    _make_paragraph(region_display, body_style),
                    _make_paragraph(severity.title(), body_style),
                    _make_paragraph(issue_title, body_style),
                    _make_paragraph(description, body_style),
                    _make_paragraph(remediation, body_style),
                ]
            )

    detail_table = Table(
        detail_data,
        colWidths=[
            0.8 * inch,
            1.1 * inch,
            0.9 * inch,
            0.8 * inch,
            0.8 * inch,
            1.1 * inch,
            1.3 * inch,
            1.3 * inch,
        ],
    )

    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("BOX", (0, 0), (-1, -1), 0.25, colors.grey),
        ]
    )

    for row_index in range(1, len(detail_data)):
        severity_cell = detail_data[row_index][4]
        if isinstance(severity_cell, Paragraph):
            severity_text = severity_cell.text.upper()
        else:
            severity_text = str(severity_cell).upper()
        severity_color = SEVERITY_COLOR_MAP.get(severity_text, colors.black)
        table_style.add("TEXTCOLOR", (3, row_index), (3, row_index), severity_color)

    detail_table.setStyle(table_style)
    flowables.append(detail_table)
    flowables.append(Spacer(1, 12))
    flowables.append(
        Paragraph(
            f"Report generated by Cloud Auditor on {generation_date_text}.",
            body_style,
        )
    )

    if scan_logs:
        flowables.append(Spacer(1, 24))
        flowables.append(Paragraph("Scan Logs", section_style))
        log_headers = ["Timestamp", "Level", "Message"]
        log_rows = [[_make_paragraph(header, body_style) for header in log_headers]]
        for log in scan_logs:
            timestamp = _format_datetime(getattr(log, "created_at", None))
            level = getattr(log, "level", "N/A")
            message = getattr(log, "message", "N/A")
            log_rows.append(
                [
                    _make_paragraph(timestamp, body_style),
                    _make_paragraph(level.title(), body_style),
                    _make_paragraph(message, body_style),
                ]
            )

        log_table = Table(
            log_rows,
            colWidths=[1.5 * inch, 1.0 * inch, 3.3 * inch],
        )
        log_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("BOX", (0, 0), (-1, -1), 0.25, colors.grey),
                ]
            )
        )
        flowables.append(log_table)

    doc.build(flowables)
    buffer.seek(0)
    return buffer


def generate_excel_report(
    findings: Iterable,
    report_context: dict,
    requested_by,
    scan_logs: Optional[Sequence] = None,
) -> io.BytesIO:
    """Return an Excel report for the provided findings."""

    findings_list = list(findings)
    generation_date = timezone.now()
    generation_date_text = _format_datetime(generation_date)
    scan_job = report_context.get("scan_job")
    scan_jobs = report_context.get("scan_jobs") or ([scan_job] if scan_job else [])
    total_findings = len(findings_list)
    scan_date_text = _format_scan_date_range(scan_jobs)
    requested_by_name = requested_by.get_full_name() or requested_by.username
    requested_by_email = requested_by.email or "n/a"
    cloud_account = report_context.get("cloud_account") or getattr(scan_job, "cloud_account", None)
    provider = cloud_account.get_provider_display() if cloud_account else "N/A"
    account_name = cloud_account.account_name if cloud_account else "N/A"

    workbook = Workbook()
    metadata_sheet = workbook.active
    metadata_sheet.title = "Report Info"
    metadata_sheet.append(["Field", "Value"])
    metadata_header_font = Font(bold=True)
    for cell in metadata_sheet[1]:
        cell.font = metadata_header_font

    scan_job_ids = [str(job.id) for job in scan_jobs if job]
    scan_job_label = scan_job_ids[0] if len(scan_job_ids) == 1 else (
        f"{scan_job_ids[0]} (+{len(scan_job_ids) - 1} more)"
        if scan_job_ids
        else "N/A"
    )
    scan_scope_value = (
        "No scans" if not scan_jobs else (
            f"Scan job #{scan_job_label}" if len(scan_jobs) == 1 else f"{len(scan_jobs)} scan jobs (latest #{scan_job_label})"
        )
    )
    scan_regions_text = _format_scan_regions(report_context.get("scan_regions"))

    metadata_entries = [
        ("Report Title", "Cloud Security Scan Report"),
        ("Scan Scope", scan_scope_value),
        ("Scan Job ID", scan_job_label),
        ("Scan Jobs", str(len(scan_jobs))),
        ("Total Findings", str(total_findings)),
        ("Cloud Account", account_name),
        ("Provider", provider),
        ("Regions Scanned", scan_regions_text),
        (
            "Requested By",
            f"{requested_by_name} ({requested_by_email})",
        ),
        ("Scan Date", scan_date_text),
        ("Report Generated", generation_date_text),
    ]
    for label, value in metadata_entries:
        metadata_sheet.append([label, value])

    data_sheet = workbook.create_sheet(title="Scan Findings")
    headers = [
        "Scan Job ID",
        "Resource ID",
        "Service",
        "Region",
        "Severity",
        "Issue Title",
        "Description",
        "Remediation",
        "Scan Date",
    ]
    data_sheet.append(headers)
    for cell in data_sheet[1]:
        cell.font = Font(bold=True)

    for finding in findings_list:
        resource_id = finding.resource_id or "N/A"
        service = getattr(finding, "resource_type", "N/A")
        job = getattr(finding, "scan_job", None)
        region = getattr(finding, "region", None)
        if not region and job:
            regions_text = job.target_regions_text()
            region = regions_text
        region_display = region or "N/A"
        severity = getattr(finding, "severity", "N/A").title()
        issue_title = getattr(finding, "issue_type", "N/A")
        description = getattr(finding, "description", "N/A")
        remediation = getattr(finding, "recommendation", "N/A")
        scan_date_value = getattr(finding, "created_at", None)
        scan_date_value_text = scan_date_value.strftime("%Y-%m-%d %H:%M:%S") if scan_date_value else "N/A"

        row = [
            str(job.id) if job else "N/A",
            resource_id,
            service,
            region_display,
            severity,
            issue_title,
            description,
            remediation,
            scan_date_value_text,
        ]
        data_sheet.append(row)

        severity_fill = PatternFill(
            start_color=EXCEL_SEVERITY_COLOR.get(severity.upper(), "FFFFFF"),
            end_color=EXCEL_SEVERITY_COLOR.get(severity.upper(), "FFFFFF"),
            fill_type="solid",
        )
        data_sheet.cell(row=data_sheet.max_row, column=5).fill = severity_fill

    data_sheet.auto_filter.ref = data_sheet.dimensions
    data_sheet.freeze_panes = "A2"

    for column_cells in data_sheet.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = min(max(max_length + 2, 15), 50)
        data_sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    if scan_logs:
        logs_sheet = workbook.create_sheet(title="Scan Logs")
        headers = ["Timestamp", "Level", "Message"]
        logs_sheet.append(headers)
        for cell in logs_sheet[1]:
            cell.font = Font(bold=True)

        for log in scan_logs:
            timestamp = _format_datetime(getattr(log, "created_at", None))
            level = getattr(log, "level", "N/A")
            message = getattr(log, "message", "N/A")
            logs_sheet.append([timestamp, level.title(), message])

        for column_cells in logs_sheet.columns:
            max_length = 0
            column = column_cells[0].column
            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = min(max(max_length + 2, 20), 60)
            logs_sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        logs_sheet.auto_filter.ref = logs_sheet.dimensions
        logs_sheet.freeze_panes = "A2"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
